import matplotlib
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from xlsx2html import xlsx2html
from jinja2 import Environment, FileSystemLoader
import pdfkit
import pandas as pd
import re
from multiprocessing import Pool
from functools import partial


data_year = [lambda x:'Год',
lambda x:'Средняя зарплата',
lambda x:'Средняя зарплата - '+x,
lambda x:'Количество вакансий',
lambda x:'Количество вакансий - '+x]

data_city = [lambda x:'Город', lambda x:'Уровень зарплат', lambda x:'', lambda x:'Город', lambda x:'Доля вакансий']

data_out = ['Динамика уровня зарплат по годам:',
'Динамика количества вакансий по годам:',
'Динамика уровня зарплат по годам для выбранной профессии:',
'Динамика количества вакансий по годам для выбранной профессии:',
'Уровень зарплат по городам (в порядке убывания):',
'Доля вакансий по городам (в порядке убывания):']

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,}

thins = Side(border_style="thin", color='000000')
style = Border(top=thins, bottom=thins, left=thins, right=thins)

class Vacancy:
    """Класс для предстваления вакансий.

    Attributes:
        name (string): Название вакансии
        description (str): Описание вакансии
        key_skills (List[str]): Ключевые навыки для вакансии
        experience_id (str): Требуемый опыт
        premium (str): Атрибут, отвечающий за премиальность вакансии
        employer_name (str): Название компании
        salary (Salary): Информация о зарплате
        area_name (str): Название города
        published_at (str): Дата публикации вакансии
    """
    def __init__(self, dict_vac: dict[str, str]):
        """
        Инициализирует объект Vacancy, проверяя наличие некоторых полей для вакансии
        Args: dict_vac (Dict[str, str]): Словарь, хранящий информацию о вакансии. Ключи - это названия полей,
        значения - информация о вакансии.
        """
        self.name = dict_vac['name']
        self.description = 'Нет данных' if 'description' not in dict_vac.keys() else dict_vac['description']
        self.key_skills = 'Нет данных' if 'key_skills' not in dict_vac.keys() else dict_vac['key_skills'].split('\n')
        self.experience_id = 'Нет данных' if 'experience_id' not in dict_vac.keys() else dict_vac['experience_id']
        self.premium = 'Нет данных' if 'premium' not in dict_vac.keys() else dict_vac['premium']
        self.employer_name = 'Нет данных' if 'employer_name' not in dict_vac.keys() else dict_vac['employer_name']
        self.salary = Salary(dict_vac['salary_from'], dict_vac['salary_to'], dict_vac['salary_currency'])
        self.area_name = dict_vac['area_name']
        self.published_at = parse_date_with_str(dict_vac['published_at'])


class Salary:
    """Класс для предстваления зарплаты.

        Attributes:
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
            current_salary (int): Средняя граница вилки оклада
        """
    def __init__(self, salary_from: str, salary_to: str, salary_currency: str):
        """Инициализирует объект Salary, создает salary (зарплата) с типом Salary.

        Args:
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.current_salary = currency_to_rub[salary_currency] * (int(float(salary_to)) + int(float(salary_from))) // 2


class Report:
    """Класс для предстваления статистики.

        Attributes:
            profession (str): Название профессии
            years (list[int]): Промежуток времени
            sr_salary (list[int]): Средние зарплаты
            salary_count (list[int]): Кол-во зарплат
            sr_prof_salary (list[int]): Средние зарплаты для данной профессии
            prof_salary_count (list[int]): Кол-во зарплат для данной профессии
            all_vacancy_count (int): Кол-во всех вакансий
            city_for_salary (list[str]): Города для статистики по заралате
            city_salary (list[int]): Заралаты по городам
            city_for_vacancy (list[str]): Города для статистики по кол-ву вакансий
            city_vacancy (list[float]): Кол-во вакансий по городам
    """
    def __init__(self, profession: str, years_all_data, years_profession, city_all_data: dict[str, tuple[int or float, int]]):
        """Инициализирует объект Report.

        :param profession: Название профессии
        :param years_all_data: Словарь, где ключ - год, значение - (зарплата, кол-во вакансий)
        :param years_profession: Словарь, где ключ - год, значение - (зарплата для данной профессии, кол-во вакансий для данной профессии)
        :param city_all_data: Словарь, где ключ - город, значение - (зарплата, кол-во вакансий)
        """
        self.profession = profession
        self.years = list(years_all_data.keys())
        self.sr_salary = [v1 for v1,v2 in years_all_data.values()]
        self.salary_count = [v2 for v1,v2 in years_all_data.values()]
        self.sr_prof_salary = [v1 for v1,v2 in years_profession.values()]
        self.prof_salary_count = [v2 for v1,v2 in years_profession.values()]
        self.all_vacancy_count = sum(self.salary_count)
        self.city_for_salary, self.city_salary, self.city_for_vacancy, self.city_vacancy = Report.find_city_key(city_all_data, self.all_vacancy_count)


    def find_city_key(dictionary: dict[str, tuple[int or float, int]], all_vacancy_count: int) -> tuple[list, list[int], list, list]:
        """Вычисляет значения для полей city_for_salary, city_salary, city_for_vacancy, city_vacancy

        :param dictionary: Словарь, где ключ - город, значение - (зарплата, кол-во вакансий)
        :param all_vacancy_count: Кол-во всех зарплат
        :return: Значения для полей city_for_salary, city_salary, city_for_vacancy, city_vacancy
        """
        res_dic1 = {}
        res_dic2 = {}
        for key, value in dictionary.items():
            if all_vacancy_count / 100 <= value[1]:
                res_dic1[key] = int(float(value[0] // value[1]))
                res_dic2[key] = round(value[1] / all_vacancy_count, 4)

        res_dic1 = {k: v for k, v in sorted(res_dic1.items(), key=lambda item: item[1], reverse=True)}
        res_dic2 = {k: v for k, v in sorted(res_dic2.items(), key=lambda item: item[1], reverse=True)}

        res_dic1 = dict(list(res_dic1.items())[:10])
        res_dic2 = dict(list(res_dic2.items())[:10])
        return (list(res_dic1.keys()),list(res_dic1.values()), list(res_dic2.keys()),list(res_dic2.values()))


    def generate_excel(self):
        """Генерирует exel таблицу

        :return: report.xlsx
        """
        wb = Workbook()
        ws1 = wb.create_sheet('Статистика по годам')
        ws2 = wb.create_sheet('Статистика по городам')
        wb.remove(wb['Sheet'])
        for i in range(len(self.years)):
            self.fill_sheet(ws1, i, 1, self.years, data_year, int)
            self.fill_sheet(ws1, i, 2, self.sr_salary, data_year, int)
            self.fill_sheet(ws1, i, 3, self.sr_prof_salary, data_year, int)
            self.fill_sheet(ws1, i, 4, self.salary_count, data_year, int)
            self.fill_sheet(ws1, i, 5, self.prof_salary_count, data_year, int)
            if i < len(self.city_for_salary):
                self.fill_sheet(ws2, i, 1, self.city_for_salary, data_city, str)
                self.fill_sheet(ws2, i, 2, self.city_salary, data_city, int)
                self.fill_sheet(ws2, i, 3, [''] * len(self.city_for_salary), data_city, str)
                self.fill_sheet(ws2, i, 4, self.city_for_vacancy, data_city, str)
                self.fill_sheet(ws2, i, 5, self.city_vacancy, data_city, float)
        for ws in wb:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value != '':
                        cell.border = style
        wb.save('report.xlsx')

    def fill_sheet(self, sheet, i: int, column: int, value, naming: str, format):
        """Заполняет листы для таблицы exel

        :param sheet: Название листа
        :param i: номер строки
        :param column: номер колонки
        :param value: Значения в столбце
        :param naming: Название столбца
        :param format: Формат поля
        :return: Заполненый лист для exel
        """
        if i == 0:
            sheet.cell(row=i + 1, column=column).value = naming[column-1](self.profession)
            sheet.cell(row=i + 1, column=column).font = Font(bold=True)
            my_value = [len(str(v)) for v in value]
            sheet.column_dimensions[get_column_letter(column)].width = max(len(naming[column - 1](self.profession)), max(my_value))+2
        sheet.cell(row=i + 2,column=column).value = format(value[i])
        if format == float:
            sheet.cell(row=i + 2, column=column).number_format = '0.00%'

    def generate_image(self):
        """Генерирует картинку графиков статистики

        :return: graph.png
        """
        matplotlib.rc('font', size=8)
        width = 0.4
        fig, ((picture1, picture2), (picture3, picture4)) = plt.subplots(nrows=2, ncols=2)

        # First image
        x = np.arange(len(self.years))
        picture1.set_title('Уровень зарплат по годам')
        picture1.bar(x - width / 2, self.sr_salary, width, label='средняя з/п')
        picture1.bar(x + width / 2, self.sr_prof_salary, width, label=f'з/п {self.profession}')
        picture1.legend(loc='upper left')
        picture1.grid(axis='y')
        picture1.set_xticks(x, self.years, rotation=90)

        # Second image
        picture2.set_title('Количество вакансий по годам')
        picture2.bar(x - width / 2, self.salary_count, width, label='Количество вакансий')
        picture2.bar(x + width / 2, self.prof_salary_count, width, label=f'Количество вакансий {self.profession}')
        picture2.legend(loc='upper left')
        picture2.grid(axis='y')
        picture2.set_xticks(x, self.years, rotation=90)

        # Third image
        cities = [city.replace(' ', '\n').replace('-', '-\n') for city in self.city_for_salary]
        y_pos = np.arange(len(cities))
        picture3.set_title('Уровень зарплат по городам')
        picture3.barh(y_pos, self.city_salary, align='center')
        picture3.invert_yaxis()
        picture3.grid(axis='x')
        picture3.set_yticks(y_pos, labels=cities, fontsize=6)

        # Fourth image
        name_labels = self.city_for_vacancy + ['Другие']
        city_percent = self.city_vacancy + [1 - sum(self.city_vacancy)]
        picture4.set_title('Доля вакансий по городам')
        picture4.pie(city_percent, labels=name_labels, radius=1.32, textprops={'fontsize': 6})

        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_dictionaries(self):
        """Выводит значения статистики в консоль

        :return: Значения статистики в консоль
        """
        dictionaries = [dict(zip(self.years, self.sr_salary)), dict(zip(self.years, self.salary_count)),
                        dict(zip(self.years, self.sr_prof_salary)), dict(zip(self.years, self.prof_salary_count)),
                        dict(zip(self.city_for_salary, self.city_salary)), dict(zip(self.city_for_vacancy, self.city_vacancy))]
        for i in range(len(dictionaries)):
            print(data_out[i], dictionaries[i])

    def generate_pdf(self):
        """Генерирует pdf файд

        :return: report.pdf
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        out1 = xlsx2html('report.xlsx', sheet='Статистика по годам')
        out1.seek(0)
        code1 = out1.read()
        out2 = xlsx2html('report.xlsx', sheet='Статистика по городам')
        out2.seek(0)
        code2 = out2.read()

        pdf_template = template.render({'profession': self.profession, 'table1': code1, 'table2': code2})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class DataSet:
    """Класс для представления вакансий

    Attributes:
        file_name (str): Название файла
        profession (str): Название профессии
        df (DataFrame) : сформированный DataFrame без пустых значений
        list_csv (list[str]) : Лист с ссылками на csv файлы
        list_vacancy (list[Vacancy]): Список всех вакансий
    """
    def __init__(self, file_name: str, profession : str, columns_for_report: list[str]):
        """Инициализирует объект DataSet, создает list_vacancy (лист вакансий) с типом list[Vacancy],
        df (DataFrame) и list_csv (Лист с ссылками на csv файлы)

        :param file_name: Название файла
        :param profession: Название профессии
        :param columns_for_report: Столбцы для составления отчета в будущем
        """
        self.file_name = file_name
        self.profession = profession
        df, list_csv, list_vacancy = self.create_df_csv_vacancy(columns_for_report)
        self.df = df
        self.list_csv = list_csv
        self.list_vacancy = list_vacancy

    def create_df_csv_vacancy(self, columns_for_report: list[str]):
        """Создает df, list_csv, list_vacancy соотвецтвенно

        :param columns_for_report: Столбцы для составления отчета
        :return: df, list_csv, list_vacancy соотвецтвенно
        """
        list_for_csv = []
        list_csv = []
        list_vacancy = []
        reader = pd.read_csv(self.file_name, on_bad_lines='skip')
        reader['published_at'] = reader['published_at'].apply(lambda s: s[:4])
        years = reader['published_at'].unique()
        column_names = reader.columns.tolist()
        for row in reader.values.tolist():
            if without_empty(row, len(column_names)):
                row_dict = dict(zip(column_names, row))
                list_vacancy.append(Vacancy(row_dict))
                vac = [row_dict[column] for column in columns_for_report]
                list_for_csv.append(vac)
        df = pd.DataFrame(list_for_csv, columns=columns_for_report)
        #df.to_csv(f'clean_csv\\clean_{self.file_name}.csv')
        for year in years:
            #data = df[df['published_at'] == year]
            #data.to_csv(f'csv_files\\year_{year}.csv')
            list_csv.append(f'csv_files\\year_{year}.csv')
        return (df, list_csv, list_vacancy)

    def find_dynamics_for_city(self):
        """Находит значения для составления статистики по городам

        :return: словарь статистики по городам (dict[str, tuple[int or float, int]])
        """
        city_all_data = {}
        for vacancy in self.list_vacancy:
            city_all_data = update(city_all_data, vacancy.area_name, vacancy.salary.current_salary)
        return city_all_data


def update(dictionary: dict, key: any, current_salary: int or float) -> dict:
    """Обнавляет значение в словаре

    :param dictionary: Словарь
    :param key: Ключ
    :param current_salary: Средняя граница вилки оклада
    :return: Обновленный словарь
    """
    if key in dictionary.keys():
        dictionary[key] = (dictionary[key][0] + current_salary, dictionary[key][1] + 1)
    else:
        dictionary[key] = (current_salary, 1)
    return dictionary


def get_statistic(df, list_csv,  profession):
    years = df['published_at'].unique()
    years_all_data = {year: [] for year in years}
    years_profession = {year: [] for year in years}
    statistics = [years_all_data, years_profession]
    p = Pool()
    output = list(p.map(partial(get_statistic_by_year, profession=profession, statistics=statistics), list_csv))
    years_all_data = {stat[0][0]: (stat[0][1], stat[0][2]) for stat in output}
    years_profession = {stat[1][0]: (stat[1][1], stat[1][2]) for stat in output}
    return (years_all_data, years_profession)


def get_statistic_by_year(csv_file, profession, statistics):
    df = pd.read_csv(csv_file)
    df['salary'] = df[['salary_from', 'salary_to']].mean(axis=1)
    year = int(csv_file[15:19])
    statistics[0] = (year, int(df['salary'].mean()), len(df))
    statistics[1] = (year, int(df[df['name'].str.contains(profession, regex=False)]['salary'].mean()), len(df[df['name'].str.contains(profession, regex=False)]))
    return statistics


def without_empty(my_list: list, count: int) -> bool:
    """Проверяет лист на наличие пустых значений

    :param my_list: Лист
    :param count: Корректное кол-во элементов списка
    :return: равенство count и my_list после удаления пустых значений из последнего
    """
    clean_list = [x for x in my_list if str(x) != 'nan']
    return len(clean_list) == count


def clean_string(raw_with_html: str) -> str:
    """
    Очищает строку от HTML кода
    Args:
        raw_html (str): Строка, которую нужно очистить
    Returns:
        str: Очищенная строка.
    """
    result = re.sub("<.*?>", '', raw_with_html)
    return result if '\n' in raw_with_html else " ".join(result.split())


def parse_date_with_str(date_vac: str) -> str:
    """Форматирует дату публикации к нужному формату 'обрезая' строку

    :param date_vac: Дата публикации
    :return: Отформатированная дата публикации
    """
    return date_vac[:4]


""""*****************************************************************************************************************"""
data_entry = ["Введите название файла: ",
"Введите название профессии: "]
columns_for_report = ['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at']

if __name__ == '__main__':
    # file_name = input(data_entry[0])
    # profession = input(data_entry[1])
    file_name = 'vacancies_by_year.csv'
    profession = 'Программист'

    dataSet = DataSet(file_name, profession, columns_for_report)
    years_all_data, years_profession = get_statistic(dataSet.df, dataSet.list_csv, profession)
    city_all_data = dataSet.find_dynamics_for_city()
    report = Report(profession, years_all_data, years_profession, city_all_data)
    report.generate_dictionaries()
    # report.generate_excel()
    # report.generate_image()
    # report.generate_pdf()


