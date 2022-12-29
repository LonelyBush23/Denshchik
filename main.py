import csv
import datetime
import matplotlib
from dateutil.parser import parse
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from xlsx2html import xlsx2html
from jinja2 import Environment, FileSystemLoader
import pdfkit
from unittest import TestCase
import dateutil


data_entry = ["Введите название файла: ",
"Введите название профессии: "]


data_year = [lambda x:'Год',
lambda x:'Средняя зарплата',
lambda x:'Средняя зарплата - '+x,
lambda x:'Количество вакансий',
lambda x:'Количество вакансий - '+x]

data_city = [lambda x:'Город', lambda x:'Уровень зарплат', lambda x:'', lambda x:'Город', lambda x:'Доля вакансий']

data_out = ['Динамика уровня зарплат по годам:',
'Динамика количества вакансий по годам:',
'Динамика уровня зарплат по годам для выбранной профессии:',
'Динамика количества вакансий по годам для выбранной профессии:',
'Уровень зарплат по городам (в порядке убывания):',
'Доля вакансий по городам (в порядке убывания):']

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,}

thins = Side(border_style="thin", color='000000')
style = Border(top=thins, bottom=thins, left=thins, right=thins)

class Vacancy:
    """Класс для предстваления вакансий.

    Attributes:
        name (str): Название вакансии
        salary (Salary): Зарплата
        area_name (str): Город
        published_at (str): Дата публикации
    """
    def __init__(self, name: str, salary_from: str, salary_to: str, salary_currency: str, area_name: str, published_at: str):
        """Инициализирует объект Vacancy, создает salary (зарплата) с типом Salary.

        Args:
            name (str): Название вакансии
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
            area_name (str): Город
            published_at (str): Дата публикации

        >>> type(Vacancy('Программист' ,'10', '60', 'RUR', 'Салда', '2007')).__name__
        'Vacancy'
        >>> Vacancy('Программист' ,'10', '60', 'RUR', 'Салда', '2007').published_at
        2007
        >>> Vacancy('Программист' ,'10', '60', 'RUR', 'Салда', '2007').name
        'Программист'
        """
        self.name = name
        self.salary = Salary(salary_from, salary_to, salary_currency)
        self.area_name = area_name
        self.published_at = int(published_at)


class Salary:
    """Класс для предстваления зарплаты.

        Attributes:
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
            current_salary (int): Средняя граница вилки оклада
        """
    def __init__(self, salary_from: str, salary_to: str, salary_currency: str):
        """Инициализирует объект Salary, создает salary (зарплата) с типом Salary.

        Args:
            salary_from (int): Нижняя граница вилки оклада
            salary_to (int): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада

        >>> type(Salary('10', '60', 'RUR')).__name__
        'Salary'
        >>> Salary('10', '60', 'RUR').current_salary
        35
        >>> Salary('10', '60', 'USD').current_salary
        2123.0
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.current_salary = currency_to_rub[salary_currency] * (int(float(salary_to)) + int(float(salary_from))) // 2


class Report:
    """Класс для предстваления статистики.

        Attributes:
            profession (str): Название профессии
            years (list[int]): Промежуток времени
            sr_salary (list[int]): Средние зарплаты
            salary_count (list[int]): Кол-во зарплат
            sr_prof_salary (list[int]): Средние зарплаты для данной профессии
            prof_salary_count (list[int]): Кол-во зарплат для данной профессии
            all_vacancy_count (int): Кол-во всех вакансий
            city_for_salary (list[str]): Города для статистики по заралате
            city_salary (list[int]): Заралаты по городам
            city_for_vacancy (list[str]): Города для статистики по кол-ву вакансий
            city_vacancy (list[float]): Кол-во вакансий по городам
    """
    def __init__(self, profession: str, years_all_data: dict[int, tuple[int or float, int]], years_profession: dict[int, tuple[int or float, int]], city_all_data: dict[str, tuple[int or float, int]]):
        """Инициализирует объект Report.

        :param profession: Название профессии
        :param years_all_data: Словарь, где ключ - год, значение - (зарплата, кол-во вакансий)
        :param years_profession: Словарь, где ключ - год, значение - (зарплата для данной профессии, кол-во вакансий для данной профессии)
        :param city_all_data: Словарь, где ключ - город, значение - (зарплата, кол-во вакансий)
        """
        self.profession = profession
        self.years = list(years_all_data.keys())
        self.sr_salary = Report.find_sal(years_all_data)
        self.salary_count = [v2 for v1,v2 in years_all_data.values()]
        self.sr_prof_salary = Report.find_sal(years_profession)
        self.prof_salary_count = [v2 for v1,v2 in years_profession.values()]
        self.all_vacancy_count = sum(self.salary_count)
        self.city_for_salary, self.city_salary, self.city_for_vacancy, self.city_vacancy = Report.find_city_key(city_all_data, self.all_vacancy_count)

    def find_sal(dictionary: dict[any, tuple[int or float, int]]) -> list[int]:
        """Вычисляет зарплату

        :return: Лист заплат
        """
        sal = []
        for v1,v2 in dictionary.values():
            try:
                sal.append(int(float(v1 // v2)))
            except:
                sal.append(v1)
        return(sal)


    def find_city_key(dictionary: dict[str, tuple[int or float, int]], all_vacancy_count: int) -> tuple[list, list[int], list, list]:
        """Вычисляет значения для полей city_for_salary, city_salary, city_for_vacancy, city_vacancy

        :param dictionary: Словарь, где ключ - город, значение - (зарплата, кол-во вакансий)
        :param all_vacancy_count: Кол-во всех зарплат
        :return: Значения для полей city_for_salary, city_salary, city_for_vacancy, city_vacancy
        """
        res_dic1 = {}
        res_dic2 = {}
        for key, value in dictionary.items():
            if all_vacancy_count / 100 <= value[1]:
                res_dic1[key] = int(float(value[0] // value[1]))
                res_dic2[key] = round(value[1] / all_vacancy_count, 4)

        res_dic1 = {k: v for k, v in sorted(res_dic1.items(), key=lambda item: item[1], reverse=True)}
        res_dic2 = {k: v for k, v in sorted(res_dic2.items(), key=lambda item: item[1], reverse=True)}

        res_dic1 = dict(list(res_dic1.items())[:10])
        res_dic2 = dict(list(res_dic2.items())[:10])
        return (list(res_dic1.keys()),list(res_dic1.values()), list(res_dic2.keys()),list(res_dic2.values()))


    def generate_excel(self):
        """Генерирует exel таблицу

        :return: report.xlsx
        """
        wb = Workbook()
        ws1 = wb.create_sheet('Статистика по годам')
        ws2 = wb.create_sheet('Статистика по городам')
        wb.remove(wb['Sheet'])
        for i in range(len(self.years)):
            self.fill_sheet(ws1, i, 1, self.years, data_year, int)
            self.fill_sheet(ws1, i, 2, self.sr_salary, data_year, int)
            self.fill_sheet(ws1, i, 3, self.sr_prof_salary, data_year, int)
            self.fill_sheet(ws1, i, 4, self.salary_count, data_year, int)
            self.fill_sheet(ws1, i, 5, self.prof_salary_count, data_year, int)
            if i < len(self.city_for_salary):
                self.fill_sheet(ws2, i, 1, self.city_for_salary, data_city, str)
                self.fill_sheet(ws2, i, 2, self.city_salary, data_city, int)
                self.fill_sheet(ws2, i, 3, [''] * len(self.city_for_salary), data_city, str)
                self.fill_sheet(ws2, i, 4, self.city_for_vacancy, data_city, str)
                self.fill_sheet(ws2, i, 5, self.city_vacancy, data_city, float)
        for ws in wb:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value != '':
                        cell.border = style
        wb.save('report.xlsx')

    def fill_sheet(self, sheet, i: int, column: int, value, naming: str, format):
        """Заполняет листы для таблицы exel

        :param sheet: Название листа
        :param i: номер строки
        :param column: номер колонки
        :param value: Значения в столбце
        :param naming: Название столбца
        :param format: Формат поля
        :return: Заполненый лист для exel
        """
        if i == 0:
            sheet.cell(row=i + 1, column=column).value = naming[column-1](self.profession)
            sheet.cell(row=i + 1, column=column).font = Font(bold=True)
            my_value = [len(str(v)) for v in value]
            sheet.column_dimensions[get_column_letter(column)].width = max(len(naming[column - 1](self.profession)), max(my_value))+2
        sheet.cell(row=i + 2,column=column).value = format(value[i])
        if format == float:
            sheet.cell(row=i + 2, column=column).number_format = '0.00%'

    def generate_image(self):
        """Генерирует картинку графиков статистики

        :return: graph.png
        """
        matplotlib.rc('font', size=8)
        width = 0.4
        fig, ((picture1, picture2), (picture3, picture4)) = plt.subplots(nrows=2, ncols=2)

        # First image
        x = np.arange(len(self.years))
        picture1.set_title('Уровень зарплат по годам')
        picture1.bar(x - width / 2, self.sr_salary, width, label='средняя з/п')
        picture1.bar(x + width / 2, self.sr_prof_salary, width, label=f'з/п {self.profession}')
        picture1.legend(loc='upper left')
        picture1.grid(axis='y')
        picture1.set_xticks(x, self.years, rotation=90)

        # Second image
        picture2.set_title('Количество вакансий по годам')
        picture2.bar(x - width / 2, self.salary_count, width, label='Количество вакансий')
        picture2.bar(x + width / 2, self.prof_salary_count, width, label=f'Количество вакансий {self.profession}')
        picture2.legend(loc='upper left')
        picture2.grid(axis='y')
        picture2.set_xticks(x, self.years, rotation=90)

        # Third image
        cities = [city.replace(' ', '\n').replace('-', '-\n') for city in self.city_for_salary]
        y_pos = np.arange(len(cities))
        picture3.set_title('Уровень зарплат по городам')
        picture3.barh(y_pos, self.city_salary, align='center')
        picture3.invert_yaxis()
        picture3.grid(axis='x')
        picture3.set_yticks(y_pos, labels=cities, fontsize=6)

        # Fourth image
        name_labels = self.city_for_vacancy + ['Другие']
        city_percent = self.city_vacancy + [1 - sum(self.city_vacancy)]
        picture4.set_title('Доля вакансий по городам')
        picture4.pie(city_percent, labels=name_labels, radius=1.32, textprops={'fontsize': 6})

        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_dictionaries(self):
        """Выводит значения статистики в консоль

        :return: Значения статистики в консоль
        """
        dictionaries = [dict(zip(self.years, self.sr_salary)), dict(zip(self.years, self.salary_count)),
                        dict(zip(self.years, self.sr_prof_salary)), dict(zip(self.years, self.prof_salary_count)),
                        dict(zip(self.city_for_salary, self.city_salary)), dict(zip(self.city_for_vacancy, self.city_vacancy))]
        for i in range(len(dictionaries)):
            print(data_out[i], dictionaries[i])

    def generate_pdf(self):
        """Генерирует pdf файд

        :return: report.pdf
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        out1 = xlsx2html('report.xlsx', sheet='Статистика по годам')
        out1.seek(0)
        code1 = out1.read()
        out2 = xlsx2html('report.xlsx', sheet='Статистика по городам')
        out2.seek(0)
        code2 = out2.read()

        pdf_template = template.render({'profession': self.profession, 'table1': code1, 'table2': code2})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

class DataSet:
    """Класс для представления вакансий.

    Attributes:
        file_name (str): Название файла
        list_vacancy (list[Vacancy]): Список всех вакансий
    """
    def __init__(self, file_name: str):
        """Инициализирует объект DataSet, создает list_vacancy (лист вакансий) с типом list[Vacancy].

        Args:
            file_name (str): Название файла
        """
        self.file_name = file_name
        self.list_vacancy = DataSet.create_list_vacancy(file_name)

    def csv_reader(file_name: str):
        """Считывает файл.

        :param file_name: Название файла
        :return: Считанный файл
        """
        file = open(file_name, encoding='utf_8_sig')
        return csv.reader(file)


    def create_list_vacancy(file_name : str) -> list[Vacancy]:
        """Создает list_vacancy (лист вакансий)

        :param file_name: Название файла
        :return: Лист вакансий
        """
        list_row = DataSet.csv_reader(file_name)
        list_vacancy = []
        is_first_row = True
        for row in list_row:
            if is_first_row:
                naming_dic = create_naming_dic(row)
                is_first_row = False
            elif without_empty(row, len(naming_dic)):
                list_vacancy.append(Vacancy(row[naming_dic.get('name')], row[naming_dic.get('salary_from')], row[naming_dic.get('salary_to')], row[naming_dic.get('salary_currency')], row[naming_dic.get('area_name')], parse_date_with_str(row[naming_dic.get('published_at')])))
        return list_vacancy


    def find_dynamics(self, profession) -> Report:
        """Находит значения для составления статистики

        :param profession: Название профессии
        :return: объект класса Report
        """
        years_all_data = {}
        years_profession = {}
        city_all_data = {}
        for vacancy in self.list_vacancy:
            years_all_data = update(years_all_data, vacancy.published_at, vacancy.salary.current_salary)
            city_all_data = update(city_all_data, vacancy.area_name, vacancy.salary.current_salary)
            if profession in vacancy.name and len(profession) != 0:
                years_profession = update(years_profession, vacancy.published_at, vacancy.salary.current_salary)
            elif vacancy.published_at not in years_profession.keys():
                years_profession[vacancy.published_at] = (0, 0)
        return Report(profession, years_all_data, years_profession, city_all_data)


def update(dictionary: dict, key: any, current_salary: int or float) -> dict:
    """Обнавляет значение в словаре

    :param dictionary: Словарь
    :param key: Ключ
    :param current_salary: Средняя граница вилки оклада
    :return: Обновленный словарь

    >>> update({'a':(2,1), 'b': (1, 1)}, 'b', 23)
    {'a': (2, 1), 'b': (24, 2)}
    >>> update({'a':(2,1), 'b': (1, 1)}, 'a', 5)
    {'a': (7, 2), 'b': (1, 1)}
    >>> update({'a':(2,1), 'b': (1, 1)}, 'c', 3)
    {'a': (2, 1), 'b': (1, 1), 'c': (3, 1)}
    """
    if key in dictionary.keys():
        dictionary[key] = (dictionary[key][0] + current_salary, dictionary[key][1] + 1)
    else:
        dictionary[key] = (current_salary, 1)
    return dictionary


def without_empty(my_list: list, count: int) -> bool:
    """Проверяет лист на наличие пустых значений

    :param my_list: Лист
    :param count: Корректное кол-во элементов списка
    :return: равенство count и my_list после удаления пустых значений из последнего
    >>> without_empty(['1','','3'], 3)
    False
    >>> without_empty(['1','7','3'], 4)
    False
    >>> without_empty(['1','7','3'], 3)
    True
    >>> without_empty(['1','','3'], 2)
    True
    """
    try:
        my_list.remove('')
    except:
        pass
    return len(my_list) == count

def create_naming_dic(naming: list[str]) -> dict[str, int]:
    """Создает словарь, где ключ это название поля, а значение - его позиция

    :param naming: Лист названий столбцов
    :return: Словарь для названия столбцов по позициям

    >>> create_naming_dic(['cat', 'car', 'star'])
    {'cat': 0, 'car': 1, 'star': 2}
    """
    result = {}
    for i in range(len(naming)):
        result[naming[i]] = i
    return result

# def parse_date_with_datetime(date_vac: str) -> str:
#     """Форматирует дату публикации к нужному формату используя datetime
#
#     :param date_vac: Дата публикации
#     :return: Отформатированная дата публикации
#     """
#     return datetime.datetime.strptime(date_vac, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')

def parse_date_with_str(date_vac: str) -> str:
    """Форматирует дату публикации к нужному формату 'обрезая' строку

    :param date_vac: Дата публикации
    :return: Отформатированная дата публикации
    """
    return date_vac[:4]

# def parse_date_with_dateutil(date_vac: str) -> str:
#     """Форматирует дату публикации к нужному формату используя dateutil
#
#     :param date_vac: Дата публикации
#     :return: Отформатированная дата публикации
#     """
#     return parse(date_vac).strftime('%Y')

""""*****************************************************************************************************************"""
# file_name = input(data_entry[0])
# profession = input(data_entry[1])
file_name = 'vacancies_by_year.csv'
profession = 'Программист'

dataSet = DataSet(file_name)

report = dataSet.find_dynamics(profession)
report.generate_dictionaries()
#report.generate_excel()
#report.generate_image()
#report.generate_pdf()

class without_empty(TestCase):
    def test1(self):
        self.assertEqual(without_empty(['1','','3'], 3), False)
    def test2(self):
        self.assertEqual(without_empty(['1','7','3'], 4), False)
    def test3(self):
        self.assertEqual(without_empty(['1','7','3'], 3), True)
    def test4(self):
        self.assertEqual(without_empty(['1','','3'], 2), True)


if __name__ == "__main__":
    import doctest
    import unittest
    #doctest.testmod()
    unittest.main()
